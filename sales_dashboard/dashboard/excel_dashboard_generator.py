import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class ExcelDashboardGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Dashboard"
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.subheader_font = Font(bold=True, size=12, color="FFFFFF")
        self.title_font = Font(bold=True, size=16, color="2E86AB")
        self.normal_font = Font(size=10)
        self.header_fill = PatternFill(
            start_color="2E86AB", end_color="2E86AB", fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="A23B72", end_color="A23B72", fill_type="solid"
        )
        self.success_fill = PatternFill(
            start_color="28A745", end_color="28A745", fill_type="solid"
        )
        self.warning_fill = PatternFill(
            start_color="FFC107", end_color="FFC107", fill_type="solid"
        )
        self.danger_fill = PatternFill(
            start_color="DC3545", end_color="DC3545", fill_type="solid"
        )
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_dashboard(self, data, predictions=None, insights=None, client_name="Client"):
        """Create a comprehensive Excel dashboard"""
        self.client_name = client_name
        self.data = data
        self.predictions = predictions or {}
        self.insights = insights or {}
        self.create_header()
        self.create_summary_section()
        self.create_kpi_section()
        self.create_data_section()
        self.create_charts_section()
        self.create_predictions_section()
        self.create_insights_section()
        self.create_recommendations_section()
        self.auto_adjust_columns()
        return self.wb

    def create_header(self):
        """Create the dashboard header"""
        self.ws['A1'] = (
            f"Targetorate - {self.client_name} Marketing Analytics Dashboard"
        )
        self.ws['A1'].font = Font(bold=True, size=20, color="2E86AB")
        self.ws.merge_cells('A1:H1')
        self.ws['A2'] = (
            f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        )
        self.ws['A2'].font = Font(italic=True, size=10, color="666666")
        self.ws.merge_cells('A2:H2')
        self.ws['A4'] = ""
        self.ws.row_dimensions[1].height = 28
        self.ws.row_dimensions[2].height = 18
        self.ws.row_dimensions[4].height = 8

    def create_summary_section(self):
        """Create summary statistics section"""
        row = 5
        self.ws[f'A{row}'] = "📊 EXECUTIVE SUMMARY"
        self.ws[f'A{row}'].font = self.title_font
        self.ws.merge_cells(f'A{row}:H{row}')
        self.ws.row_dimensions[row].height = 22
        row += 2
        if not self.data.empty:
            total_spend = self.data.get('spend', pd.Series([0])).sum()
            total_revenue = self.data.get('revenue', pd.Series([0])).sum()
            total_clicks = self.data.get('clicks', pd.Series([0])).sum()
            total_impressions = self.data.get('impressions', pd.Series([0])).sum()
            total_conversions = self.data.get('conversions', pd.Series([0])).sum()
            roas = total_revenue / total_spend if total_spend > 0 else 0
            ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
            conversion_rate = (total_conversions / total_clicks * 100) if total_clicks > 0 else 0
            cpa = total_spend / total_conversions if total_conversions > 0 else 0
            summary_data = [
                ["Metric", "Value", "Status"],
                ["Total Spend", f"${total_spend:,.2f}", "💰"],
                ["Total Revenue", f"${total_revenue:,.2f}", "💵"],
                ["ROAS", f"{roas:.2f}x", "📈" if roas > 1 else "📉"],
                ["Click-Through Rate", f"{ctr:.2f}%", "✅" if ctr > 2 else "⚠️"],
                ["Conversion Rate", f"{conversion_rate:.2f}%", "✅" if conversion_rate > 2 else "⚠️"],
                ["Cost Per Acquisition", f"${cpa:.2f}", "✅" if cpa < 50 else "⚠️"],
                ["Total Clicks", f"{total_clicks:,}", "📊"],
                ["Total Impressions", f"{total_impressions:,}", "👁️"],
                ["Total Conversions", f"{total_conversions:,}", "🎯"]
            ]
            for i, row_data in enumerate(summary_data):
                for j, cell_value in enumerate(row_data):
                    cell = self.ws.cell(row=row + i, column=j + 1, value=cell_value)
                    if i == 0:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.font = self.normal_font
                        cell.border = self.border
                        if j == 0:
                            cell.font = Font(bold=True, size=10)
                        elif j == 2:
                            cell.alignment = Alignment(horizontal='center')
                        elif j == 1:
                            cell.alignment = Alignment(horizontal='right')
                self.ws.row_dimensions[row + i].height = 18
            self.ws.freeze_panes = self.ws['A7']
            row += len(summary_data) + 2
        self.ws[f'A{row}'] = ""
        self.ws.row_dimensions[row].height = 8
        return row + 1

    def create_kpi_section(self):
        """Create KPI cards section"""
        row = 20
        self.ws[f'A{row}'] = "🎯 KEY PERFORMANCE INDICATORS"
        self.ws[f'A{row}'].font = self.title_font
        self.ws.merge_cells(f'A{row}:H{row}')
        row += 2
        if not self.data.empty:
            platforms = self.data.get('platform', pd.Series(['Unknown'])).unique()
            for i, platform in enumerate(platforms):
                platform_data = self.data[
                    self.data.get('platform', pd.Series(['Unknown'])) == platform
                ]
                if not platform_data.empty:
                    spend = platform_data.get('spend', pd.Series([0])).sum()
                    revenue = platform_data.get('revenue', pd.Series([0])).sum()
                    clicks = platform_data.get('clicks', pd.Series([0])).sum()
                    impressions = platform_data.get('impressions', pd.Series([0])).sum()
                    roas = revenue / spend if spend > 0 else 0
                    ctr = (clicks / impressions * 100) if impressions > 0 else 0
                    col_offset = (i % 2) * 4
                    start_col = 1 + col_offset
                    self.ws.cell(row=row, column=start_col, value=f"📱 {platform.upper()}")
                    self.ws.cell(row=row, column=start_col).font = self.subheader_font
                    self.ws.cell(row=row, column=start_col).fill = self.subheader_fill
                    self.ws.merge_cells(
                        f'{chr(64 + start_col)}{row}:{chr(64 + start_col + 3)}{row}'
                    )
                    row += 1
                    kpis = [
                        ["Spend", f"${spend:,.2f}"],
                        ["Revenue", f"${revenue:,.2f}"],
                        ["ROAS", f"{roas:.2f}x"],
                        ["CTR", f"{ctr:.2f}%"]
                    ]
                    for j, (kpi_name, kpi_value) in enumerate(kpis):
                        self.ws.cell(row=row + j, column=start_col, value=kpi_name).font = Font(bold=True)
                        self.ws.cell(row=row + j, column=start_col + 1, value=kpi_value)
                        self.ws.cell(row=row + j, column=start_col + 2, value="✅" if roas > 1 else "⚠️")
                        for col in range(start_col, start_col + 3):
                            self.ws.cell(row=row + j, column=col).border = self.border
                    row += len(kpis) + 1
                    if i % 2 == 1:
                        row += 2
        return row + 2

    def create_data_section(self):
        """Create raw data section"""
        row = 35
        self.ws[f'A{row}'] = "📋 RAW DATA"
        self.ws[f'A{row}'].font = self.title_font
        self.ws.merge_cells(f'A{row}:H{row}')
        self.ws.row_dimensions[row].height = 22
        row += 2
        if not self.data.empty:
            for r_idx, r in enumerate(dataframe_to_rows(self.data.head(100), index=False, header=True)):
                for j, value in enumerate(r, 1):
                    cell = self.ws.cell(row=row + r_idx, column=j, value=value)
                    if r_idx == 0:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.font = self.normal_font
                        cell.border = self.border
                        if r_idx % 2 == 1:
                            cell.fill = PatternFill(
                                start_color="F7F7F7", end_color="F7F7F7", fill_type="solid"
                            )
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal='center')
                        elif isinstance(value, str) and value.startswith('$'):
                            cell.alignment = Alignment(horizontal='right')
                self.ws.row_dimensions[row + r_idx].height = 16
            self.ws.auto_filter.ref = f"A{row}:H{row + r_idx}"
            self.ws.freeze_panes = self.ws[f'A{row + 1}']
            row += r_idx + 1
            if len(self.data) > 100:
                self.ws[f'A{row}'] = (
                    f"Note: Showing first 100 rows of {len(self.data)} total records"
                )

    def create_charts_section(self):
        """Create charts section"""
        pass

    def create_predictions_section(self):
        """Create predictions section"""
        pass

    def create_insights_section(self):
        """Create insights section"""
        pass

    def create_recommendations_section(self):
        """Create recommendations section"""
        pass

    def auto_adjust_columns(self):
        """Auto-adjust column widths"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

